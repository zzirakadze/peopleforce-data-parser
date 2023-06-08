import calendar
import logging
import os
import time
from datetime import date
from typing import Dict, Any
import pprint
import chromedriver_autoinstaller
import openpyxl
import psycopg2
from dotenv import load_dotenv
from pydantic import BaseSettings
from selenium import webdriver
from selenium.common import (
    ElementNotInteractableException,
    ElementClickInterceptedException,
    NoSuchElementException,
    TimeoutException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

######################################## LOGGING ########################################

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler()],
)

######################################## CONFIG SETUP ########################################

load_dotenv()


class CONFIG(BaseSettings):
    DATABASE_HOST: str
    DATABASE_NAME: str
    DATABASE_USERNAME: str
    DATABASE_PASSWORD: str
    EMAIL: str
    PASS: str
    DOMAIN_URL: str
    REPORTS_URL: str

    class Config:
        env_file = ".env"
        env_file_encoding = "utf-8"


######################################## DATABASE #################################################
class DatabaseCursor(psycopg2.extensions.cursor):
    def __init__(self, connection):
        super().__init__(connection)

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.connection.close()


class Database:
    def __init__(self):
        self.conn = None

    def connect(self) -> None:
        if self.conn is not None:
            return

        self.conn = psycopg2.connect(
            host=CONFIG().DATABASE_HOST,
            database=CONFIG().DATABASE_NAME,
            user=CONFIG().DATABASE_USERNAME,
            password=CONFIG().DATABASE_PASSWORD,
        )

    def close(self):
        if self.conn is not None:
            self.conn.close()

    def get_session(self) -> DatabaseCursor:
        self.connect()
        return DatabaseCursor(self.conn)


######################################## SELENIUM FUNCTIONS ########################################
def wait_and_input(
    driver: webdriver,
    locator_with_selector: tuple,
    input_text: str,
    wait_time: int = 10,
) -> None:
    """
    :param driver: webdriver instance
    :param locator_with_selector: takes locator with selector as a tuple eg: (By.XPATH, "//button[@type='submit']")
    :param input_text: text which you want to input
    :return:
    """
    web_element: WebElement = WebDriverWait(driver, wait_time).until(
        EC.presence_of_element_located(locator_with_selector)
    )
    try:
        web_element.send_keys(input_text)
    except ElementNotInteractableException as e:
        logging.error(e)


def wait_and_click(driver: webdriver, locator_with_selector: tuple) -> None:
    """
    :param locator_with_selector : takes locator with selector as a tuple eg: (By.XPATH, "//button[@type='submit']")
    :return: returns none
    """
    web_element: WebElement = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(locator_with_selector)
    )
    try:
        web_element.click()
    except ElementClickInterceptedException as e:
        logging.error(e)


def find_elements(driver: webdriver, locator_with_selector: tuple) -> list:
    """
    :param locator_with_selector: takes locator with selector as a tuple eg: (By.XPATH, "//button[@type='submit']")
    :return: returns list of web elements
    """
    try:
        return WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(locator_with_selector)
        )
    except NoSuchElementException as e:
        logging.error(e)


def element_is_visible(driver: webdriver, locator_with_selector: tuple) -> bool:
    """
    :param locator_with_selector: takes locator with selector as a tuple eg: (By.XPATH, "//button[@type='submit']")
    :return: returns bool
    """
    try:
        WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located(locator_with_selector)
        )
        logging.info(f"Element with selector {locator_with_selector[1]} is visible")
        return True
    except TimeoutException as e:
        logging.error(
            f"Element with selector {locator_with_selector[1]} is not visible", e
        )


######################################## Driver instanciation ########################################
class Driver:
    @classmethod
    def instance(cls):
        chromedriver_autoinstaller.install()
        download_directory = os.getcwd()
        prefs = {
            "download.default_directory": download_directory,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options = webdriver.ChromeOptions()
        options.add_experimental_option("prefs", prefs)
        options.add_argument("--start-maximized")
        # options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-features=EnableEphemeralFlashPermission")
        browser = webdriver.Chrome(options=options)
        return browser

    def __init__(self):
        self.instance = self.instance()


######################################## Page Objects ########################################
class Authorize:
    email_field: tuple[By, str] = (By.ID, "user_email")
    password_field: tuple[By, str] = (By.ID, "user_password")
    login_button: tuple[By, str] = (By.XPATH, "//button[@type='submit']")
    profile_actions: tuple[By, str] = (By.XPATH, "//a[@data-cy='profile_actions']")

    def __init__(self, driver):
        self.driver = driver

    def login(self) -> None:
        self.driver.get(CONFIG().DOMAIN_URL)

        self.input(self.email_field, CONFIG().EMAIL).input(
            self.password_field, CONFIG().PASS
        ).click(self.login_button).wait_for_visible(self.profile_actions)

    def input(self, field, value) -> "Authorize":
        wait_and_input(self.driver, field, value)
        return self

    def click(self, field) -> "Authorize":
        wait_and_click(self.driver, field)
        return self

    def wait_for_visible(self, field) -> "Authorize":
        assert element_is_visible(
            self.driver, field
        ), f"Login failed, {field[1]} is not visible"
        return self


class GetExcel:
    dropdown: tuple[By, str] = (
        By.XPATH,
        "//div[@class='col-auto']/a[@data-bs-toggle='dropdown']",
    )
    excel: tuple[By, str] = (By.XPATH, "//a[contains(text(),'Export to Excel')]")

    def __init__(self, driver) -> None:
        self.driver = driver

    def redirect_to_reports_page(self) -> None:
        self.driver.get(CONFIG().REPORTS_URL)

    def download_excel(self) -> None:
        self.redirect_to_reports_page()
        download_directory = os.getcwd()
        filename = "ppf_data.xlsx"
        try:
            wait_and_click(self.driver, self.dropdown)
            wait_and_click(self.driver, self.excel)
            time.sleep(2)
            fname = [fname for fname in os.listdir() if fname.endswith(".xlsx")]
            wait_time = 0
            while not any(fname.endswith(".xlsx") for fname in os.listdir()):
                if wait_time > 10:
                    raise TimeoutException("File not downloaded")
                time.sleep(1)
                wait_time += 1
            os.rename(
                download_directory + "/" + fname[0], download_directory + "/" + filename
            )
        except Exception as e:
            logging.error("Failed to download excel", e)
        finally:
            self.driver.quit()


######################################## Data Sorter ########################################
class DataSorter:
    user_data = {}

    def transfer_data(self, file_path: str) -> dict[Any, Any]:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook["Sheet1"]
        _, max_days = calendar.monthrange(date.today().year, date.today().month)
        for row in range(2, worksheet.max_row + 1):
            name_surname = (
                worksheet.cell(row=row, column=2).value
                + " "
                + worksheet.cell(row=row, column=4).value
            )
            time_offs = {}
            for day in range(1, max_days + 1):
                value = worksheet.cell(row=row, column=day + 12).value
                if value in [
                    "Vacation",
                    "Sick Leave",
                    "Work from Home (WFH)",
                    "Unpaid Day Off",
                    "Vacation (Georgia)",
                ]:
                    time_offs.setdefault(value, []).append(day)
            if not time_offs:
                self.user_data[name_surname] = [{"No Time Offs": []}]
            else:
                self.user_data[name_surname] = [{k: v} for k, v in time_offs.items()]
        return self.user_data


class SearchBySpecificUser:
    def __init__(self, user_data: dict[Any, Any]):
        self.user_data = user_data

    def search(self, user: str) -> dict[Any, Any]:
        return {user: self.user_data[user]}


######################################## CRAWLER ########################################
if __name__ == "__main__":

    def crawler():
        driver = Driver().instance
        Authorize(driver).login()
        GetExcel(driver).download_excel()
        filename = "ppf_data.xlsx"
        data = DataSorter().transfer_data(filename)
        print(data)

    crawler()
